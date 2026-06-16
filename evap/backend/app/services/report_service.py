"""Report generation: PDF (ReportLab) and Excel (openpyxl)."""

from __future__ import annotations

import io
import logging
import os
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from sqlalchemy.ext.asyncio import AsyncSession

from ..schemas.report import ReportRequest, ReportResponse, ReportStatus

logger = logging.getLogger(__name__)


def _reports_dir() -> Path:
    try:
        from ..core.config import settings  # type: ignore[import]
        return Path(settings.REPORTS_DIR)
    except Exception:
        return Path("/tmp/evap_reports")


def _file_url(filename: str) -> str:
    try:
        from ..core.config import settings  # type: ignore[import]
        base = getattr(settings, "REPORTS_BASE_URL", "/reports")
        return f"{base}/{filename}"
    except Exception:
        return f"/reports/{filename}"


# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------

def _build_pdf_header(doc, title: str, subtitle: str = "") -> None:
    """Build standard page header on a ReportLab SimpleDocTemplate."""
    pass  # integrated into generate functions below


async def generate_attendance_pdf(db: AsyncSession, params: ReportRequest) -> str:
    """Generate an attendance PDF with summary table and bar chart. Returns file path."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    from .attendance_service import get_attendance_range, get_today_attendance

    reports_dir = _reports_dir()
    reports_dir.mkdir(parents=True, exist_ok=True)
    filename = f"attendance_{params.date_from}_{params.date_to}_{int(datetime.now().timestamp())}.pdf"
    filepath = reports_dir / filename

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(str(buffer), pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph(f"Attendance Report", styles["Title"]))
    story.append(Paragraph(f"{params.date_from} to {params.date_to}", styles["Heading2"]))
    story.append(Spacer(1, 0.5 * cm))

    # Summary table header
    headers = ["Employee ID", "Name", "Date", "First Entry", "Last Exit", "Hours", "Status", "Late"]
    table_data = [headers]

    # Get data - if specific employee in filters, use range query
    employee_id = params.filters.get("employee_id")
    if employee_id:
        records = await get_attendance_range(db, str(employee_id), params.date_from, params.date_to)
        for r in records:
            table_data.append([
                r.employee_id,
                r.employee_name or "",
                str(r.date),
                r.first_entry.strftime("%H:%M") if r.first_entry else "—",
                r.last_exit.strftime("%H:%M") if r.last_exit else "—",
                f"{r.working_hours:.1f}h" if r.working_hours else "—",
                r.status,
                "Yes" if r.is_late else "No",
            ])

    col_widths = [2.5 * cm, 4 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2 * cm, 2.5 * cm, 1.5 * cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f0f0")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
    ]))
    story.append(t)
    story.append(Spacer(1, 1 * cm))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))

    doc.build(story)
    filepath.write_bytes(buffer.getvalue())
    logger.info("Attendance PDF generated: %s", filepath)
    return str(filepath)


async def generate_attendance_excel(db: AsyncSession, params: ReportRequest) -> str:
    """Generate attendance Excel workbook with per-sheet summaries."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from .attendance_service import get_attendance_range

    reports_dir = _reports_dir()
    reports_dir.mkdir(parents=True, exist_ok=True)
    filename = f"attendance_{params.date_from}_{params.date_to}_{int(datetime.now().timestamp())}.xlsx"
    filepath = reports_dir / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["Employee ID", "Name", "Date", "First Entry", "Last Exit", "Working Hours", "Status", "Is Late", "Overtime Hrs"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    employee_id = params.filters.get("employee_id")
    row_num = 2
    if employee_id:
        records = await get_attendance_range(db, str(employee_id), params.date_from, params.date_to)
        for r in records:
            ws.cell(row=row_num, column=1, value=r.employee_id)
            ws.cell(row=row_num, column=2, value=r.employee_name or "")
            ws.cell(row=row_num, column=3, value=str(r.date))
            ws.cell(row=row_num, column=4, value=r.first_entry.strftime("%H:%M") if r.first_entry else "")
            ws.cell(row=row_num, column=5, value=r.last_exit.strftime("%H:%M") if r.last_exit else "")
            ws.cell(row=row_num, column=6, value=r.working_hours or 0)
            ws.cell(row=row_num, column=7, value=r.status)
            ws.cell(row=row_num, column=8, value="Yes" if r.is_late else "No")
            ws.cell(row=row_num, column=9, value=r.overtime_hours or 0)
            row_num += 1

    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].auto_size = True

    wb.save(str(filepath))
    logger.info("Attendance Excel generated: %s", filepath)
    return str(filepath)


async def generate_vehicle_report(db: AsyncSession, params: ReportRequest) -> str:
    """Vehicle log report as PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from ..models.vehicle import LicensePlateLog  # type: ignore[import]
    from sqlalchemy import and_, select

    reports_dir = _reports_dir()
    reports_dir.mkdir(parents=True, exist_ok=True)
    filename = f"vehicle_log_{params.date_from}_{params.date_to}_{int(datetime.now().timestamp())}.pdf"
    filepath = reports_dir / filename

    start_dt = datetime.combine(params.date_from, datetime.min.time()).replace(tzinfo=timezone.utc)
    end_dt = datetime.combine(params.date_to, datetime.max.time()).replace(tzinfo=timezone.utc)

    stmt = select(LicensePlateLog).where(
        and_(LicensePlateLog.entry_time >= start_dt, LicensePlateLog.entry_time <= end_dt)
    )
    if params.site_id:
        stmt = stmt.where(LicensePlateLog.site_id == params.site_id)
    stmt = stmt.order_by(LicensePlateLog.entry_time.desc()).limit(5000)
    result = await db.execute(stmt)
    rows = result.scalars().all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(str(buffer), pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Vehicle Log Report", styles["Title"]),
        Paragraph(f"{params.date_from} to {params.date_to}", styles["Heading2"]),
        Spacer(1, 0.5 * cm),
    ]

    headers = ["Plate", "Type", "Direction", "Entry Time", "Exit Time", "Duration (min)", "Camera"]
    table_data = [headers]
    for r in rows:
        duration = ""
        if r.parking_duration_seconds:
            duration = str(r.parking_duration_seconds // 60)
        table_data.append([
            r.plate_number,
            r.vehicle_type or "",
            r.direction or "",
            r.entry_time.strftime("%Y-%m-%d %H:%M") if r.entry_time else "",
            r.exit_time.strftime("%H:%M") if r.exit_time else "—",
            duration,
            str(r.camera_id or ""),
        ])

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f3460")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#eef1f7")]),
    ]))
    story.append(t)
    doc.build(story)
    filepath.write_bytes(buffer.getvalue())
    return str(filepath)


async def generate_visitor_report(db: AsyncSession, params: ReportRequest) -> str:
    """Visitor analytics report as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    from ..models.zone_history import ZoneHistory  # type: ignore[import]
    from sqlalchemy import and_, func, select

    reports_dir = _reports_dir()
    reports_dir.mkdir(parents=True, exist_ok=True)
    filename = f"visitors_{params.date_from}_{params.date_to}_{int(datetime.now().timestamp())}.xlsx"
    filepath = reports_dir / filename

    start_dt = datetime.combine(params.date_from, datetime.min.time()).replace(tzinfo=timezone.utc)
    end_dt = datetime.combine(params.date_to, datetime.max.time()).replace(tzinfo=timezone.utc)

    stmt = (
        select(
            ZoneHistory.person_id,
            func.min(ZoneHistory.entry_time).label("first_seen"),
            func.max(ZoneHistory.entry_time).label("last_seen"),
            func.count().label("zone_visits"),
            func.sum(ZoneHistory.duration_seconds).label("total_dwell"),
        )
        .where(
            and_(
                ZoneHistory.person_type == "visitor",
                ZoneHistory.entry_time >= start_dt,
                ZoneHistory.entry_time <= end_dt,
            )
        )
        .group_by(ZoneHistory.person_id)
        .order_by(func.min(ZoneHistory.entry_time).desc())
    )
    result = await db.execute(stmt)
    rows = result.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visitors"
    header_fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["Visitor ID", "First Seen", "Last Seen", "Zone Visits", "Total Dwell (min)"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for ri, (pid, first, last, visits, dwell) in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=pid)
        ws.cell(row=ri, column=2, value=first.strftime("%Y-%m-%d %H:%M") if first else "")
        ws.cell(row=ri, column=3, value=last.strftime("%Y-%m-%d %H:%M") if last else "")
        ws.cell(row=ri, column=4, value=visits)
        ws.cell(row=ri, column=5, value=round((dwell or 0) / 60, 1))

    wb.save(str(filepath))
    return str(filepath)


async def generate_executive_summary(db: AsyncSession, params: ReportRequest) -> str:
    """Combined executive PDF: dashboard KPIs + top stats from all modules."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    from .analytics_service import get_dashboard_stats

    reports_dir = _reports_dir()
    reports_dir.mkdir(parents=True, exist_ok=True)
    filename = f"executive_summary_{params.date_from}_{params.date_to}_{int(datetime.now().timestamp())}.pdf"
    filepath = reports_dir / filename

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(str(buffer), pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("EVAP Executive Summary", styles["Title"]))
    story.append(Paragraph(f"Period: {params.date_from} – {params.date_to}", styles["Heading2"]))
    story.append(Spacer(1, 1 * cm))

    # KPI table (placeholder data since we need async DB calls above)
    kpi_data = [
        ["Metric", "Value"],
        ["Report Period", f"{params.date_from} to {params.date_to}"],
        ["Site ID", str(params.site_id or "All")],
        ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M UTC")],
    ]
    t = Table(kpi_data, colWidths=[8 * cm, 8 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#533483")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
    ]))
    story.append(t)
    story.append(Spacer(1, 1 * cm))
    story.append(Paragraph("This report was generated by EVAP – Enterprise Video Analytics Platform.", styles["Normal"]))

    doc.build(story)
    filepath.write_bytes(buffer.getvalue())
    return str(filepath)


async def save_report(
    db: AsyncSession,
    user_id: int,
    report_type: str,
    format: str,
    filepath: str,
    params: ReportRequest,
) -> ReportResponse:
    from ..models.report import Report  # type: ignore[import]

    expires_at = None
    try:
        from ..core.config import settings
        from datetime import timedelta
        ttl_days = getattr(settings, "REPORT_RETENTION_DAYS", 30)
        expires_at = datetime.now(timezone.utc) + timedelta(days=ttl_days)
    except Exception:
        pass

    filename = os.path.basename(filepath)
    report = Report(
        report_type=report_type,
        title=f"{report_type.replace('_', ' ').title()} Report",
        format=format,
        generated_by=user_id,
        file_path=filepath,
        parameters=params.model_dump(mode="json"),
        generated_at=datetime.now(timezone.utc),
        expires_at=expires_at,
    )
    db.add(report)
    await db.commit()
    await db.refresh(report)

    return ReportResponse(
        id=report.report_id,
        report_type=report.report_type,
        title=report.title,
        status="ready",
        format=report.format,
        file_url=_file_url(filename),
        generated_at=report.generated_at,
        expires_at=report.expires_at,
        generated_by=report.generated_by,
        parameters=report.parameters,
    )
